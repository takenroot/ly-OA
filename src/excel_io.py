"""每天新建一个 xlsx:基于参考模板复制样式,写入当日所有记录."""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from .validate import (
    COL_序号, COL_装车日期, COL_装车时间, COL_货品, COL_车牌, COL_司机,
    COL_皮重, COL_毛重, COL_净重, COL_备注, Record,
)


def _dt_to_excel_time(dt: datetime) -> float:
    """datetime → Excel 时间小数."""
    t = dt.time()
    return (t.hour * 3600 + t.minute * 60 + t.second) / 86400.0


def date_to_filename(date: datetime) -> str:
    """把日期转成 固废填埋日统计表{MMDD}.xlsx."""
    return f"固废填埋日统计表{date.month:02d}{date.day:02d}.xlsx"


def _path_to_hyperlink(path: Path) -> str:
    """把本地图片路径转成 Excel/WPS 可点击的本地路径.

    之前使用 file:// URI 并对中文进行 URL 编码,在 WPS 中点击时经常
    提示"无法打开指定的文件".改为未经编码的本地绝对路径(Windows 反斜杠),
    WPS 和 Excel 都能正常识别并调用系统默认的图片查看器."""
    return str(path.resolve())


def create_daily_xlsx(
    records: list[Record],
    yellow_marks_per_row: list[dict[str, str]],
    output_path: Path,
    reference_path: Path,
    image_paths: list[Path] | None = None,
    workshop_name: str = "",
) -> None:
    """每天新建一个 xlsx.

    1. 复制参考 xlsx → output_path(包含样式 + 表头 + 总行结构)
    2. 清掉参考里 row 3+ 的数据(但保留 row 1/2 表头和最后总行)
    3. 写入 records 到 row 3 开始
    4. 备注列写入照片超链接
    5. 更新总行的 SUM 公式
    6. 按车间名更新标题(A1)
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(reference_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.active  # 第一个 sheet,通常是 Sheet1

    # 按数据源目录名(车间名)更新标题
    if workshop_name:
        title_cell = ws.cell(row=1, column=1)
        if title_cell.value and isinstance(title_cell.value, str):
            title_cell.value = title_cell.value.replace("示例车间A", workshop_name)

    # 找最后一行(参考里的总行"合计"在 row 34 左右)
    # 清掉 row 3 ~ (total_row - 1) 的数据(保留表头和总行)
    total_row = None
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        # 找含"合计"的行;或最后一行(若没有"合计")
        cv = ws.cell(row=r, column=1).value
        if isinstance(cv, str) and "合计" in cv:
            total_row = r
            break
    if total_row is None:
        # 没有"合计"行,假设最后一行就是
        total_row = ws.max_row

    # openpyxl 的 insert_rows 不会自动移动合并单元格,
    # 先取消所有合并区域,插入行后再按需重新合并
    merged_ranges = list(ws.merged_cells.ranges)
    for mr in merged_ranges:
        ws.unmerge_cells(str(mr))

    # 如果记录数超过模板预留的数据行数,在合计行前插入足够行
    needed_rows = len(records)
    available_rows = total_row - 3
    if needed_rows > available_rows:
        rows_to_insert = needed_rows - available_rows
        ws.insert_rows(total_row, rows_to_insert)
        total_row += rows_to_insert

    # 清掉 row 3 ~ total_row-1 的数据(保留样式);跳过合并单元格
    for r in range(3, total_row):
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # 写入新数据
    for i, (rec, yellows) in enumerate(zip(records, yellow_marks_per_row)):
        r = 3 + i
        cell = ws.cell(row=r, column=1)
        if not isinstance(cell, MergedCell):
            cell.value = i + 1  # 序号
        if rec.装车日期:
            cell = ws.cell(row=r, column=2)
            if not isinstance(cell, MergedCell):
                cell.value = rec.装车日期.date()
            cell = ws.cell(row=r, column=3)
            if not isinstance(cell, MergedCell):
                cell.value = _dt_to_excel_time(rec.装车日期)
        cell = ws.cell(row=r, column=4)
        if not isinstance(cell, MergedCell):
            cell.value = rec.货品名称
        cell = ws.cell(row=r, column=5)
        if not isinstance(cell, MergedCell):
            cell.value = rec.车牌号
        cell = ws.cell(row=r, column=6)
        if not isinstance(cell, MergedCell):
            cell.value = rec.司机姓名
        if rec.皮重_吨 is not None:
            cell = ws.cell(row=r, column=7)
            if not isinstance(cell, MergedCell):
                cell.value = round(rec.皮重_吨, 2)
        if rec.毛重_吨 is not None:
            cell = ws.cell(row=r, column=8)
            if not isinstance(cell, MergedCell):
                cell.value = round(rec.毛重_吨, 2)
        if rec.净重_吨 is not None:
            cell = ws.cell(row=r, column=9)
            if not isinstance(cell, MergedCell):
                cell.value = round(rec.净重_吨, 2)

        # 备注列(J)写入编号/失败提示 + 照片超链接,方便人工复核时直接点开原图
        if image_paths and i < len(image_paths):
            img_path = image_paths[i]
            cell = ws.cell(row=r, column=10)
            if not isinstance(cell, MergedCell):
                if COL_备注 in yellows:
                    base_text = "识别失败，查看照片"
                elif rec.备注:
                    base_text = f"{rec.备注} | 查看照片"
                else:
                    base_text = "查看照片"
                cell.value = base_text
                cell.hyperlink = _path_to_hyperlink(img_path)

        # 标黄
        for col_letter, reason in yellows.items():
            cell = ws[f"{col_letter}{r}"]
            if not isinstance(cell, MergedCell):
                from openpyxl.styles import PatternFill
                cell.fill = PatternFill("solid", fgColor="FFFF00")

    # 更新总行(用新数据范围)
    if total_row > 3:
        last_data_row = 3 + len(records) - 1
        cell = ws.cell(row=total_row, column=2)
        if not isinstance(cell, MergedCell):
            cell.value = f"{len(records)}条记录"
        cell = ws.cell(row=total_row, column=7)
        if not isinstance(cell, MergedCell):
            cell.value = f"=SUM(G3:G{last_data_row})"
        cell = ws.cell(row=total_row, column=8)
        if not isinstance(cell, MergedCell):
            cell.value = f"=SUM(H3:H{last_data_row})"
        cell = ws.cell(row=total_row, column=9)
        if not isinstance(cell, MergedCell):
            cell.value = f"=SUM(I3:I{last_data_row})"

    # 重新合并标题行和合计行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    if total_row > 3:
        ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=6)

    wb.save(output_path)


def append_to_default_sheet(
    xlsx_path: Path, pr, day_label: str | None = None
) -> tuple[int, str]:
    """保留原行为(调试用):往现有 xlsx 的 Sheet1 追加一行."""
    wb = load_workbook(xlsx_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    end_row = 3
    while ws.cell(row=end_row, column=1).value is not None:
        end_row += 1
    new_row = end_row
    rec = pr.record
    last_serial = ws.cell(row=end_row - 1, column=1).value if end_row > 3 else 0
    ws.cell(row=new_row, column=1, value=(last_serial or 0) + 1)
    if rec.装车日期:
        ws.cell(row=new_row, column=2, value=rec.装车日期.date())
        ws.cell(row=new_row, column=3, value=_dt_to_excel_time(rec.装车日期))
    ws.cell(row=new_row, column=4, value=rec.货品名称)
    ws.cell(row=new_row, column=5, value=rec.车牌号)
    ws.cell(row=new_row, column=6, value=rec.司机姓名)
    if rec.皮重_吨 is not None:
        ws.cell(row=new_row, column=7, value=round(rec.皮重_吨, 2))
    if rec.毛重_吨 is not None:
        ws.cell(row=new_row, column=8, value=round(rec.毛重_吨, 2))
    if rec.净重_吨 is not None:
        ws.cell(row=new_row, column=9, value=round(rec.净重_吨, 2))
    for col_letter in pr.yellow_marks:
        from openpyxl.styles import PatternFill
        ws[f"{col_letter}{new_row}"].fill = PatternFill("solid", fgColor="FFFF00")
    total_row = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == "合计":
                total_row = cell.row
                break
        if total_row:
            break
    if total_row:
        ws.cell(row=total_row, column=9, value=f"=SUM(I3:I{new_row})")
    wb.save(xlsx_path)
    return new_row, "Sheet1"
